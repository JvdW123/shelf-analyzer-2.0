"""
Shelf Analyzer 2.0 — Excel Generator Module

This module converts Claude's JSON response into a formatted Excel file with:
- 32 columns in exact schema order
- Header row with blue background and white text
- Alternating row colors
- Excel formula for Price per Liter (EUR)
- Conditional formatting for Confidence Score and Stock Status
- Auto-adjusted column widths
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import COLUMN_SCHEMA, EXCEL_CONFIG


def generate_excel(skus: list[dict], metadata: dict) -> bytes:
    """
    Generate a formatted Excel file from SKU data.
    
    Parameters:
    - skus: List of dictionaries from Claude's JSON response (one dict per SKU)
    - metadata: Dictionary with user-provided metadata (country, city, retailer, 
                store_format, store_name, shelf_location, currency)
    
    Returns:
    - bytes: Excel file content as bytes (ready for download)
    """
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = EXCEL_CONFIG["sheet_name"]
    
    # ==============================================================================
    # STEP 1: WRITE HEADER ROW
    # ==============================================================================
    
    # Extract column names from COLUMN_SCHEMA
    header_row = [col["name"] for col in COLUMN_SCHEMA]
    ws.append(header_row)
    
    # Style the header row
    header_fill = PatternFill(
        start_color=EXCEL_CONFIG["header_bg_color"],
        end_color=EXCEL_CONFIG["header_bg_color"],
        fill_type="solid"
    )
    header_font = Font(
        name=EXCEL_CONFIG["font_name"],
        size=EXCEL_CONFIG["font_size"],
        bold=True,
        color=EXCEL_CONFIG["header_font_color"]
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Set header row height
    ws.row_dimensions[1].height = EXCEL_CONFIG["header_row_height"]
    
    # Freeze the header row (row 1 stays visible when scrolling)
    ws.freeze_panes = "A2"
    
    # ==============================================================================
    # STEP 2: WRITE DATA ROWS
    # ==============================================================================
    
    # Define thin border style for all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define alternating row fill
    alt_fill = PatternFill(
        start_color=EXCEL_CONFIG["alt_row_color"],
        end_color=EXCEL_CONFIG["alt_row_color"],
        fill_type="solid"
    )
    
    # Define data font
    data_font = Font(
        name=EXCEL_CONFIG["font_name"],
        size=EXCEL_CONFIG["font_size"]
    )
    
    # Process each SKU
    for sku_index, sku in enumerate(skus):
        row_data = []
        current_row = sku_index + 2  # Data starts at row 2 (row 1 is header)
        
        # Build row data for each column in COLUMN_SCHEMA
        for col_index, col in enumerate(COLUMN_SCHEMA):
            key = col["key"]
            
            # USER-PROVIDED COLUMNS: Get value from metadata dict, NOT from Claude's JSON
            # This ensures consistency even if Claude returns slightly different values
            if key in ["country", "city", "retailer", "store_format", "store_name", 
                       "shelf_location", "currency"]:
                value = metadata.get(key, "")
            else:
                # AI-PROVIDED COLUMNS: Get value from Claude's JSON
                value = sku.get(key)
            
            # Handle None/null values gracefully
            if value is None:
                value = ""
            
            row_data.append(value)
        
        # Append the row to the worksheet
        ws.append(row_data)
        
        # Apply styling to each cell in the row
        for col_index, cell in enumerate(ws[current_row], start=1):
            # Apply font
            cell.font = data_font
            
            # Apply border
            cell.border = thin_border
            
            # Apply alternating row color (even rows get grey background)
            if current_row % 2 == 0:
                cell.fill = alt_fill
            
            # Disable text wrapping
            cell.alignment = Alignment(wrap_text=False, vertical="center")
        
        # Set row height
        ws.row_dimensions[current_row].height = EXCEL_CONFIG["data_row_height"]
        
        # ==============================================================================
        # SPECIAL HANDLING: Column 21 (Price per Liter EUR) — EXCEL FORMULA
        # ==============================================================================
        # Column 21 is index 20 (zero-based), which is column U in Excel
        # Formula: =IFERROR(S{row}/(T{row}/1000),"")
        # S = column 19 (Price EUR), T = column 20 (Packaging Size ml)
        
        price_per_liter_col_index = 20  # Zero-based index for column 21
        price_eur_col_letter = get_column_letter(19)  # Column S (Price EUR)
        packaging_size_col_letter = get_column_letter(20)  # Column T (Packaging Size ml)
        
        # Build the formula
        formula = f'=IFERROR({price_eur_col_letter}{current_row}/({packaging_size_col_letter}{current_row}/1000),"")'
        
        # Write the formula to the cell (column U)
        price_per_liter_cell = ws.cell(row=current_row, column=price_per_liter_col_index + 1)
        price_per_liter_cell.value = formula
    
    # ==============================================================================
    # STEP 3: CONDITIONAL FORMATTING
    # ==============================================================================
    
    # Apply conditional formatting to all data rows (starting from row 2)
    for row_index in range(2, len(skus) + 2):
        
        # --- CONFIDENCE SCORE (Column 32, column AF in Excel) ---
        confidence_col_index = 32  # Column AF
        confidence_cell = ws.cell(row=row_index, column=confidence_col_index)
        confidence_value = confidence_cell.value
        
        # Apply color based on confidence score thresholds
        if confidence_value is not None and isinstance(confidence_value, (int, float)):
            if confidence_value >= EXCEL_CONFIG["confidence_high"]["min"]:
                # High confidence: green fill + dark green text
                confidence_cell.fill = PatternFill(
                    start_color=EXCEL_CONFIG["confidence_high"]["bg"],
                    end_color=EXCEL_CONFIG["confidence_high"]["bg"],
                    fill_type="solid"
                )
                confidence_cell.font = Font(
                    name=EXCEL_CONFIG["font_name"],
                    size=EXCEL_CONFIG["font_size"],
                    color=EXCEL_CONFIG["confidence_high"]["font"]
                )
            elif confidence_value >= EXCEL_CONFIG["confidence_mid"]["min"]:
                # Mid confidence: yellow fill + dark yellow text
                confidence_cell.fill = PatternFill(
                    start_color=EXCEL_CONFIG["confidence_mid"]["bg"],
                    end_color=EXCEL_CONFIG["confidence_mid"]["bg"],
                    fill_type="solid"
                )
                confidence_cell.font = Font(
                    name=EXCEL_CONFIG["font_name"],
                    size=EXCEL_CONFIG["font_size"],
                    color=EXCEL_CONFIG["confidence_mid"]["font"]
                )
            else:
                # Low confidence: red fill + dark red text
                confidence_cell.fill = PatternFill(
                    start_color=EXCEL_CONFIG["confidence_low"]["bg"],
                    end_color=EXCEL_CONFIG["confidence_low"]["bg"],
                    fill_type="solid"
                )
                confidence_cell.font = Font(
                    name=EXCEL_CONFIG["font_name"],
                    size=EXCEL_CONFIG["font_size"],
                    color=EXCEL_CONFIG["confidence_low"]["font"]
                )
        
        # --- STOCK STATUS (Column 29, column AC in Excel) ---
        stock_status_col_index = 29  # Column AC
        stock_status_cell = ws.cell(row=row_index, column=stock_status_col_index)
        stock_status_value = stock_status_cell.value
        
        # Apply red formatting if "Out of Stock"
        if stock_status_value == "Out of Stock":
            stock_status_cell.fill = PatternFill(
                start_color=EXCEL_CONFIG["out_of_stock"]["bg"],
                end_color=EXCEL_CONFIG["out_of_stock"]["bg"],
                fill_type="solid"
            )
            stock_status_cell.font = Font(
                name=EXCEL_CONFIG["font_name"],
                size=EXCEL_CONFIG["font_size"],
                color=EXCEL_CONFIG["out_of_stock"]["font"],
                bold=True
            )
    
    # ==============================================================================
    # STEP 4: AUTO-ADJUST COLUMN WIDTHS
    # ==============================================================================
    
    # Define reasonable default widths based on column content type
    # These widths are in Excel character units (approximate)
    column_widths = {
        "Country": 15,
        "City": 15,
        "Retailer": 20,
        "Store Format": 18,
        "Store Name": 20,
        "Photo": 25,
        "Shelf Location": 25,
        "Shelf Levels": 12,
        "Shelf Level": 12,
        "Product Type": 15,
        "Branded/Private Label": 20,
        "Brand": 18,
        "Sub-brand": 18,
        "Product Name": 30,
        "Flavor": 25,
        "Facings": 10,
        "Price (Local Currency)": 15,
        "Currency": 10,
        "Price (EUR)": 12,
        "Packaging Size (ml)": 18,
        "Price per Liter (EUR)": 18,
        "Need State": 15,
        "Juice Extraction Method": 22,
        "Processing Method": 18,
        "HPP Treatment": 12,
        "Packaging Type": 18,
        "Claims": 30,
        "Bonus/Promotions": 25,
        "Stock Status": 15,
        "Est. Linear Meters": 18,
        "Fridge Number": 15,
        "Confidence Score": 16
    }
    
    # Apply column widths
    for col_index, col in enumerate(COLUMN_SCHEMA, start=1):
        col_letter = get_column_letter(col_index)
        col_name = col["name"]
        width = column_widths.get(col_name, 15)  # Default to 15 if not specified
        ws.column_dimensions[col_letter].width = width
    
    # ==============================================================================
    # STEP 5: SAVE TO BYTES AND RETURN
    # ==============================================================================
    
    # Save workbook to a BytesIO buffer (in-memory, not to disk)
    buffer = io.BytesIO()
    wb.save(buffer)
    
    # Return the bytes content
    return buffer.getvalue()
